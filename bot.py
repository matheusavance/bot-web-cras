
import json
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles import PatternFill
from webdriver_auto_update.chrome_app_utils import ChromeAppUtils
from webdriver_auto_update.webdriver_manager import WebDriverManager
from botcity.web import WebBot, Browser, By
from botcity.plugins.excel import BotExcelPlugin

# Instância WebBot
bot = WebBot()

# Métodos
def autoupdate_chromedriver():
    """
    Baixa/atualiza o arquivo chromedriver baseado na versão do chrome da máquina que está rodando o código.
    """

    # Usando ChromeAppUtils para inspecionar a versão do Chrome
    chrome_app_utils = ChromeAppUtils()
    chrome_app_version = chrome_app_utils.get_chrome_version()

    # Diretório de destino para armazenar o chromedriver
    driver_directory = r"C:\Users\Usuário\Desktop\code\python"

    # Instância WebDriverManager
    driver_manager = WebDriverManager(driver_directory)

    # Executa método main para gerenciar o chromedriver
    driver_manager.main()

def pesquisa_cras(cidade, nome_estado):
    """
    Preenche o campo de pesquisa do maps com o texto 'CRAS + CIDADE + ESTADO'.\n
    :param: cidade
    :param: nome_estado
    """

    bot.browse("https://www.google.com/maps")
    bot.find_element('searchboxinput', By.ID).send_keys(f'CRAS {cidade} {nome_estado} ') # {cidade} {nome_estado} 
    bot.enter()

def extracao_pesquisa_unica(id_cras, cidade, nome_estado, path_planilha):
    """
    Extrai os dados do único resultado obtido na pesquisa.\n
    :param: id_cras
    :param: path_planilha\n
    @return id_cras
    """

    # Armazena o nome do CRAS
    nome = bot.execute_javascript('return document.getElementsByClassName("DUwDvf")[0].textContent')
    
    # Checa se a div com as informações do card existe
    div_botoes_informacoes = bot.execute_javascript(f'return document.getElementsByClassName("yx21af lLU2pe XDi3Bc")[0]')
    if not div_botoes_informacoes:
        # Adota valores padrões para o caso de não haver avaliações e cria uma lista com os dados coletados
        endereco = "Sem endereço" 
        telefone = "Sem telefone" 
        link_maps = "Sem link"
        quantidade_avaliacoes = "Não existem avaliações"
        nota_cras = "Sem nota" 
        lista_variaveis_folha_cras = [id_cras, nome, cidade, nome_estado, endereco, telefone, link_maps, quantidade_avaliacoes, nota_cras]
        
        # Preenche a folha 'CRAS' da planilha com os dados extraídos
        id_cras = preenche_folha_cras(lista_variaveis_folha_cras, path_planilha)
        id_cras += 1
        return id_cras  
    
    # Armazena o endereço do CRAS
    endereco = bot.execute_javascript('return document.getElementsByClassName("Io6YTe fontBodyMedium kR99db")[0].textContent')

    # Armazena o telefone do CRAS
    quantidade_dados_card = bot.execute_javascript('return document.getElementsByClassName("CsEnBe").length')
    for i in range(quantidade_dados_card):
        texto_botao_card = bot.execute_javascript(f'return document.getElementsByClassName("CsEnBe")[{i}].ariaLabel')
        if 'Telefone' in texto_botao_card:
            telefone = texto_botao_card
            break
        else:
            telefone = 'Sem telefone'

    # Armazena o link do maps do CRAS
    bot.execute_javascript('document.getElementsByClassName("m6QErb Pf6ghf XiKgde ecceSd tLjsW ")[0].lastChild.childNodes[0].click()')
    link_maps = bot.find_element('/html/body/div[1]/div[3]/div[1]/div/div[2]/div/div[2]/div/div/div/div[3]/div[2]/div[2]/input', By.XPATH)
    bot.wait_for_element_visibility(link_maps)
    link_maps = bot.execute_javascript('return document.getElementsByClassName("vrsrZe")[0].value')
    bot.find_element('/html/body/div[1]/div[3]/div[1]/div/div[2]/div/button', By.XPATH).click()

    # Clica em "Avaliações"
    div_botoes_informacoes = bot.execute_javascript(f'return document.getElementsByClassName("yx21af lLU2pe XDi3Bc")[0]')
    if div_botoes_informacoes:
        texto_botao_avaliacoes = bot.execute_javascript(f'return document.getElementsByClassName("LRkQ2")[1].textContent')
        if texto_botao_avaliacoes == 'Avaliações':
            bot.execute_javascript(f'document.getElementsByClassName("LRkQ2")[1].click()')
            bot.wait(2000)

            # Scrolldown para atualizar o número de comentários
            altura_atual_div_card = bot.execute_javascript('return document.getElementsByClassName("m6QErb DxyBCb kA9KIf dS8AEf XiKgde ")[0].scrollHeight')
            while True:
                bot.execute_javascript('lista = document.getElementsByClassName("m6QErb DxyBCb kA9KIf dS8AEf XiKgde ")[0]; document.getElementsByClassName("m6QErb DxyBCb kA9KIf dS8AEf XiKgde ")[0].scrollTo(0, lista.scrollHeight)')
                bot.wait(2000)
                nova_altura_div_card = bot.execute_javascript('return document.getElementsByClassName("m6QErb DxyBCb kA9KIf dS8AEf XiKgde ")[0].scrollHeight')
                if nova_altura_div_card == altura_atual_div_card:
                    break
                altura_atual_div_card = nova_altura_div_card

            # Armazena a nota e a quantidades de avaliações do CRAS
            quantidade_avaliacoes = bot.find_element('//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div[2]/div/div[2]/div[3]', By.XPATH).text
            nota_cras = bot.execute_javascript(f'return document.getElementsByClassName("fontDisplayLarge")[0].textContent')

            # Cria uma lista com os dados coletados
            lista_variaveis_folha_cras = [id_cras, nome, cidade, nome_estado, endereco, telefone, link_maps, quantidade_avaliacoes, nota_cras]

            # Preenche a folha 'CRAS' da planilha com os dados extraídos
            id_cras = preenche_folha_cras(lista_variaveis_folha_cras, path_planilha)
     
            # Extrai dados e preenche a folha 'COMENTÁRIOS' da planilha
            id_cras = extrai_dados_comentarios_cras(id_cras, nome, path_planilha)
            return id_cras

        else:
            # Adota valores padrões para o caso de não haver avaliações e cria uma lista com os dados coletados
            quantidade_avaliacoes="Não existem avaliações"
            nota_cras="Sem nota"   
            lista_variaveis_folha_cras = [id_cras, nome, cidade, nome_estado, endereco, telefone, link_maps, quantidade_avaliacoes, nota_cras]
            
            # Preenche a folha 'CRAS' da planilha com os dados extraídos
            id_cras = preenche_folha_cras(lista_variaveis_folha_cras, path_planilha)
            id_cras += 1
            return id_cras
  
    else:
        # Adota valores padrões para o caso de não haver avaliações e cria uma lista com os dados coletados
        quantidade_avaliacoes="Não existem avaliações"
        nota_cras="Sem nota"   
        lista_variaveis_folha_cras = [id_cras, nome, cidade, nome_estado, endereco, telefone, link_maps, quantidade_avaliacoes, nota_cras]
        
        # Preenche a folha 'CRAS' da planilha com os dados extraídos
        id_cras = preenche_folha_cras(lista_variaveis_folha_cras, path_planilha)
        id_cras += 1
        return id_cras

def extrai_dados_cras(id_cras, cidade, nome_estado, path_planilha):
    """
    Coleta os dados de cada cras achado na pesquisa do maps.\n
    :param path_planilha\n
    @return id_cras
    """

    # Verifica se o resultado da pesquisa no maps retornou "Correspondência parcial" ou "Correspondências parciais"
    resultado_pesquisa = bot.find_element('/html/body/div[1]/div[3]/div[8]/div[9]/div/div/div[1]/div[2]/div/div[1]/div/div/div[1]/div[1]/div[1]', By.XPATH, waiting_time=2000)
    if resultado_pesquisa:
        if 'parcial' in resultado_pesquisa.text or 'parciais' in resultado_pesquisa.text:
            endereco = "Sem endereço" 
            telefone = "Sem telefone" 
            link_maps = "Sem link"
            quantidade_avaliacoes = "Não existem avaliações"
            nota_cras = "Sem nota"
            nome = "Sem nome"
            lista_variaveis_folha_cras = [id_cras, nome, cidade, nome_estado, endereco, telefone, link_maps, quantidade_avaliacoes, nota_cras]
            
            # Preenche a folha 'CRAS' da planilha com os dados extraídos
            id_cras = preenche_folha_cras(lista_variaveis_folha_cras, path_planilha)
            id_cras += 1
            return id_cras 

    # Verifica se o resultado da pesquisa nop maps retornou um único card e extrai os dados do mesmo
    div_img_resultado_pesquisa = bot.find_element('ZKCDEc', By.CLASS_NAME)
    if div_img_resultado_pesquisa:
        id_cras = extracao_pesquisa_unica(id_cras, cidade, nome_estado, path_planilha)
        return id_cras
    
    # Verifica se o resultado da pesquisa no maps retornou mais de um card
    div_cards = bot.execute_javascript('return document.getElementsByClassName("m6QErb DxyBCb kA9KIf dS8AEf XiKgde ecceSd")[1]')
    resultado_pesquisa = bot.find_element('//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[1]/div[1]/div[2]/div[2]/div[1]/h1', By.XPATH, waiting_time=2000)

    if div_cards or 'Resultados' in resultado_pesquisa.text:
        # Scrolldown para atualizar o número de cards, caso o resultado sejam vários cards
        altura_atual_div_card = bot.execute_javascript("return document.getElementsByClassName('m6QErb DxyBCb kA9KIf dS8AEf XiKgde ecceSd')[1].scrollHeight")

        while True:
            bot.execute_javascript('lista = document.getElementsByClassName("m6QErb DxyBCb kA9KIf dS8AEf XiKgde ecceSd")[1]; document.getElementsByClassName("m6QErb DxyBCb kA9KIf dS8AEf XiKgde ecceSd")[1].scrollTo(0, lista.scrollHeight)')
            bot.wait(2000)
            nova_altura_div_card = bot.execute_javascript("return document.getElementsByClassName('m6QErb DxyBCb kA9KIf dS8AEf XiKgde ecceSd')[1].scrollHeight")
            if nova_altura_div_card == altura_atual_div_card:
                break
            altura_atual_div_card = nova_altura_div_card
    
        # Extrai dados de cada card do CRAS
        quantidade_cras = bot.execute_javascript('return document.getElementsByClassName("hfpxzc").length')
        for item_cras in range(quantidade_cras):
            # Clica no card do CRAS
            bot.execute_javascript(f'document.getElementsByClassName("hfpxzc")[{item_cras}].click()')
            bot.wait(2000)

            # Armazena o nome do CRAS
            nome = bot.execute_javascript('return document.getElementsByClassName("DUwDvf")[0].textContent')

            # Checa se a div com as informações do card existe
            div_botoes_informacoes = bot.execute_javascript(f'return document.getElementsByClassName("yx21af lLU2pe XDi3Bc")[0]')
            if not div_botoes_informacoes:
                # Adota valores padrões para o caso de não haver avaliações e cria uma lista com os dados coletados
                endereco = "Sem endereço" 
                telefone = "Sem telefone" 
                link_maps = "Sem link"
                quantidade_avaliacoes = "Não existem avaliações"
                nota_cras = "Sem nota" 
                lista_variaveis_folha_cras = [id_cras, nome, cidade, nome_estado, endereco, telefone, link_maps, quantidade_avaliacoes, nota_cras]
                
                # Preenche a folha 'CRAS' da planilha com os dados extraídos
                id_cras = preenche_folha_cras(lista_variaveis_folha_cras, path_planilha)
                id_cras += 1
                return id_cras    

            # Armazena o endereço do CRAS
            endereco = bot.execute_javascript('return document.getElementsByClassName("Io6YTe fontBodyMedium kR99db")[0].textContent')

            # Armazena o telefone do CRAS
            quantidade_dados_card = bot.execute_javascript('return document.getElementsByClassName("CsEnBe").length')
            for i in range(quantidade_dados_card):
                texto_botao_card = bot.execute_javascript(f'return document.getElementsByClassName("CsEnBe")[{i}].ariaLabel')
                if 'Telefone' in texto_botao_card:
                    telefone = texto_botao_card
                    break
                else:
                    telefone = 'Sem telefone'
                
            # Armazena o link do maps do CRAS
            bot.execute_javascript('document.getElementsByClassName("m6QErb Pf6ghf XiKgde ecceSd tLjsW ")[0].lastChild.childNodes[0].click()')
            link_maps = bot.find_element('/html/body/div[1]/div[3]/div[1]/div/div[2]/div/div[2]/div/div/div/div[3]/div[2]/div[2]/input', By.XPATH)
            bot.wait_for_element_visibility(link_maps)
            link_maps = bot.execute_javascript('return document.getElementsByClassName("vrsrZe")[0].value')
            bot.find_element('/html/body/div[1]/div[3]/div[1]/div/div[2]/div/button', By.XPATH).click()

            # Clica em "Avaliações"
            div_botoes_informacoes = bot.execute_javascript(f'return document.getElementsByClassName("yx21af lLU2pe XDi3Bc")[0]')
            if div_botoes_informacoes:
                texto_botao_avaliacoes = bot.execute_javascript(f'return document.getElementsByClassName("LRkQ2")[1].textContent')
                if texto_botao_avaliacoes == 'Avaliações':
                    bot.execute_javascript(f'document.getElementsByClassName("LRkQ2")[1].click()')
                    bot.wait(2000)

                    # Scrolldown para atualizar o número de comentários
                    altura_atual_div_card = bot.execute_javascript('return document.getElementsByClassName("m6QErb DxyBCb kA9KIf dS8AEf XiKgde ")[2].scrollHeight')
                    while True:
                        bot.execute_javascript('lista = document.getElementsByClassName("m6QErb DxyBCb kA9KIf dS8AEf XiKgde ")[2]; document.getElementsByClassName("m6QErb DxyBCb kA9KIf dS8AEf XiKgde ")[2].scrollTo(0, lista.scrollHeight)')
                        bot.wait(2000)
                        nova_altura_div_card = bot.execute_javascript('return document.getElementsByClassName("m6QErb DxyBCb kA9KIf dS8AEf XiKgde ")[2].scrollHeight')
                        if nova_altura_div_card == altura_atual_div_card:
                            break
                        altura_atual_div_card = nova_altura_div_card

                else:
                    # Adota valores padrões para o caso de não haver avaliações e cria uma lista com os dados coletados
                    quantidade_avaliacoes = "Não existem avaliações"
                    nota_cras = "Sem nota"   
                    lista_variaveis_folha_cras = [id_cras, nome, cidade, nome_estado, endereco, telefone, link_maps, quantidade_avaliacoes, nota_cras]
                    
                    # Preenche a folha 'CRAS' da planilha com os dados extraídos
                    id_cras = preenche_folha_cras(lista_variaveis_folha_cras, path_planilha)
                    id_cras += 1
                    return id_cras  
                
            else:
                # Adota valores padrões para o caso de não haver avaliações e cria uma lista com os dados coletados
                quantidade_avaliacoes = "Não existem avaliações"
                nota_cras = "Sem nota"   
                lista_variaveis_folha_cras = [id_cras, nome, cidade, nome_estado, endereco, telefone, link_maps, quantidade_avaliacoes, nota_cras]
                
                # Preenche a folha 'CRAS' da planilha com os dados extraídos
                id_cras = preenche_folha_cras(lista_variaveis_folha_cras, path_planilha)
                id_cras += 1
                return id_cras
            
            # Armazena a nota e a quantidades de avaliações do CRAS
            quantidade_avaliacoes = bot.find_element('/html/body/div[1]/div[3]/div[8]/div[9]/div/div/div[1]/div[3]/div/div[1]/div/div/div[3]/div[2]/div/div[2]/div[3]', By.XPATH).text
            nota_cras = bot.execute_javascript(f'return document.getElementsByClassName("fontDisplayLarge")[0].textContent')
            
            # Cria uma lista com os dados coletados
            lista_variaveis_folha_cras = [id_cras, nome, cidade, nome_estado, endereco, telefone, link_maps, quantidade_avaliacoes, nota_cras]

            # Preenche a folha 'CRAS' da planilha com os dados extraídos
            id_cras = preenche_folha_cras(lista_variaveis_folha_cras, path_planilha)

            # Extrai dados e preenche a folha 'COMENTÁRIOS' da planilha
            id_cras = extrai_dados_comentarios_cras(id_cras, nome, path_planilha)
    
    return id_cras

def extrai_dados_comentarios_cras(id_cras, nome, path_planilha):
    """
    Coleta os dados dos comentários de cada cras achado na pesquisa do maps.\n
    :param: id_cras
    :param: nome
    :param: path_planilha\n
    @return id_cras
    """

    # Armazena os dados sobre os comentários
    quantidade_comentarios = bot.execute_javascript('return document.getElementsByClassName("al6Kxe").length')
    contador_childElements = 0
    for comentario in range(quantidade_comentarios):
        id_comentario = bot.execute_javascript(f'return document.getElementsByClassName("d4r55")[{comentario}].textContent')
        data_comentario = bot.execute_javascript(f'return document.getElementsByClassName("rsqaWe")[{comentario}].textContent')
        nota_comentario = bot.execute_javascript(f'return document.getElementsByClassName("kvMYJc")[{comentario}].ariaLabel')
        
        quantidade_comentarios_usuario = bot.execute_javascript(f'return document.getElementsByClassName("al6Kxe")[{comentario}].childElementCount')
        if quantidade_comentarios_usuario == 1: 
            quantidade_comentarios_usuario = 'Nenhuma avaliação'
        else:
            quantidade_comentarios_usuario = bot.execute_javascript(f'return document.getElementsByClassName("al6Kxe")[{comentario}].children[1].textContent')
        
        try:            
            texto_comentario = bot.execute_javascript(f'return document.getElementsByClassName("wiI7pd")[{comentario}].textContent')
        except:
            texto_comentario = 'Não existem comentários'

        lista_variaveis_folha_comentarios = [id_cras, nome, id_comentario, data_comentario, nota_comentario, quantidade_comentarios_usuario, texto_comentario]

        id_cras = preenche_folha_comentarios(lista_variaveis_folha_comentarios, path_planilha)
        contador_childElements += 1

    # Incrementa a variável id_cras
    id_cras += 1

    return id_cras

def preenche_folha_cras(lista_variaveis_folha_cras, path_planilha):
    """
    Preenche a folha 'CRAS' da planilha.\n
    :param lista_variaveis_folha_cras
    :param path_planilha\n
    @return id_cras
    """
    
    # Lê a folha 'CRAS'
    planilha_resultado = BotExcelPlugin('CRAS').read(path_planilha)

    # Cria a lista com a lista de variáveis
    lista_preenche_folha_cras = []
    lista_preenche_folha_cras.append(lista_variaveis_folha_cras)

    # Ativa a folha 'CRAS'
    lista_folhas_planilha = planilha_resultado.list_sheets()
    folha_cras = lista_folhas_planilha[0]

    # Preenche a próxima linha da folha 'CRAS' com os valores obtidos na lista
    planilha_resultado.add_rows(lista_preenche_folha_cras, sheet=folha_cras)
    planilha_resultado.write(path_planilha)

    id_cras = lista_variaveis_folha_cras[0]
    return id_cras
 
def preenche_folha_comentarios(lista_variaveis_folha_comentarios, path_planilha):
    """
    Preenche a folha 'COMENTÁRIOS' da planilha.\n
    :param lista_variaveis_folha_comentarios
    :param path_planilha\n
    @return id_cras 
    """
    
    # Lê a folha 'COMENTÁRIOS'
    planilha_resultado = BotExcelPlugin('COMENTÁRIOS').read(path_planilha)

    # Cria a lista com a lista de variáveis
    lista_preenche_folha_comentarios = []
    lista_preenche_folha_comentarios.append(lista_variaveis_folha_comentarios)

    # Ativa a folha 'COMENTÁRIOS'
    lista_folhas_planilha = planilha_resultado.list_sheets()
    folha_comentarios = lista_folhas_planilha[1]

    # Preenche a próxima linha da folha 'CRAS' com os valores obtidos na lista
    planilha_resultado.add_rows(lista_preenche_folha_comentarios, sheet=folha_comentarios)
    planilha_resultado.write(path_planilha)

    id_cras = lista_variaveis_folha_comentarios[0]
    return id_cras

def estiliza_planilha(path_planilha):
    """
    Estiliza as folhas da planilha passada.\n
    :param path_planilha
    """
    # Abre e carrega a planilha 
    planilha = load_workbook(filename=path_planilha)
    quantidade_folhas_planilha = len(planilha.sheetnames)

    # Define o estilo usados
    fundo_preto = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    texto_branco_negrito = Font(color="FFFFFF", bold=True)
    borda_fina = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Aplica os estilos células que possuem conteúdo
    for numero_folha in range(quantidade_folhas_planilha):
        planilha.active = numero_folha
        folha = planilha.active

        # Header
        for celula in folha[1]:
            celula.fill = fundo_preto
            celula.font = texto_branco_negrito

        # Células
        for linha in folha.rows:
            for celula in linha:
                if celula.value is not None:
                    celula.border = borda_fina

    planilha.save(path_planilha)

def registra_data_horario_atual(path_planilha):
    # Registra a data e horário que a extração de dados foi feita
    data_horario = datetime.datetime.now()
    data_horario_extracao = f'{data_horario.day}/{data_horario.month}/{data_horario.year} - {data_horario.hour}:{data_horario.minute}'

    # Preenche data e horário na folha 'CRAS'
    planilha_resultado = BotExcelPlugin('CRAS').read(path_planilha)
    planilha_resultado.set_cell('J', 2, data_horario_extracao)
    planilha_resultado.write(path_planilha)

# Atualiza/baixa o chromedriver, caso necessário
autoupdate_chromedriver()

def main():
    # Modo Headless
    bot.headless = False

    # Navegador usado no processo
    bot.browser = Browser.CHROME

    # Path chromedriver
    bot.driver_path = r"C:\Users\Usuário\Desktop\code\python\chromedriver.exe"

    # Path planilha 'Resultado'
    path_planilha = r"C:\Users\Usuário\Desktop\code\python\bots\bot-web-cras\planilhas\Resultado.xlsx"

    # Path arquivo JSON
    path_arquivo_json = r"C:\Users\Usuário\Desktop\code\python\bots\bot-web-cras\estados_cidades_json\estados_cidades_sp.json"

    # Pesquisa/extrai dados do CRAS da cidade analisada e preenche a planilha 'Resultado'
    with open(path_arquivo_json, 'r', encoding='utf-8') as arquivo_json:
        dados = json.load(arquivo_json)

    # Armazena o número de linhas da planilha
    planilha_resultado = BotExcelPlugin('CRAS').read(path_planilha)
    lista_planilha = planilha_resultado.as_list()
    numero_linhas_planilha = len(lista_planilha)

    # Determina o valor da variável id_cras de acordo com o número de linhas da planilha
    if numero_linhas_planilha == 1:
        id_cras = 1
    else:
        id_cras = numero_linhas_planilha

    for estado in dados['estados']:
        nome_estado = estado['nome']
        for cidade in estado['cidades']:

            print('------------------------------------') 
            print(f'CIDADE: {cidade}')    
            print(f'ESTADO: {nome_estado}')
            print('------------------------------------')    

            pesquisa_cras(cidade, nome_estado)

            extrai_dados_cras(id_cras, cidade, nome_estado, path_planilha)
        
    registra_data_horario_atual(path_planilha)

    estiliza_planilha(path_planilha)

if __name__ == '__main__':
    main()
